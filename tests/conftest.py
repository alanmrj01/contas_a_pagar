"""Fallback de teste somente para ambientes sem python-calamine.

A aplicação em produção instala python-calamine==0.5.3. Este fallback permite
executar a suíte de validação estrutural em ambientes Linux offline do pacote de
entrega, lendo apenas os XLSX de teste via openpyxl.
"""
from __future__ import annotations

import sys
import types

try:
    import python_calamine  # noqa: F401
except ImportError:
    import openpyxl

    class _Sheet:
        def __init__(self, ws):
            self.ws = ws

        def to_python(self):
            return [list(row) for row in self.ws.iter_rows(values_only=True)]

    class _Workbook:
        def __init__(self, wb):
            self._wb = wb
            self.sheet_names = list(wb.sheetnames)

        @classmethod
        def from_path(cls, path: str):
            return cls(openpyxl.load_workbook(path, read_only=True, data_only=True))

        def get_sheet_by_name(self, name: str):
            return _Sheet(self._wb[name])

    module = types.ModuleType("python_calamine")
    module.CalamineWorkbook = _Workbook
    sys.modules["python_calamine"] = module
